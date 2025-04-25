from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.urls import reverse
from django.db.models import Q
from django.http import JsonResponse
from .models import Client, Service, Building, BuildingTracking
import pandas as pd
import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
import logging
import openpyxl
import tempfile
import os

logger = logging.getLogger(__name__)

# Helper Functions
def is_admin_user(user):
    """Check if user is authenticated and has admin username."""
    return user.is_authenticated and user.username == 'admin'

def to_date(value):
    """Convert Excel date to Python date."""
    if pd.isna(value):
        return None
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.date() if isinstance(value, datetime.datetime) else value
    try:
        return pd.to_datetime(value).date()
    except:
        return None

# Excel Processing
def unmerge_excel_cells(workbook, sheet_name="LL84- Benchmarking"):
    """Unmerge cells in Excel sheet and fill with top-left value."""
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    worksheet = workbook[sheet_name]
    for merged_range in list(worksheet.merged_cells.ranges):
        top_left_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
        value = top_left_cell.value
        worksheet.unmerge_cells(
            start_row=merged_range.min_row,
            start_column=merged_range.min_col,
            end_row=merged_range.max_row,
            end_column=merged_range.max_col
        )
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                worksheet.cell(row=row, column=col).value = value
    return workbook

def save_temp_excel(file_stream):
    """Save uploaded Excel file to temporary location."""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        for chunk in file_stream.chunks():
            temp_file.write(chunk)
        return temp_file.name

def cleanup_temp_files(*file_paths):
    """Remove temporary files."""
    for path in file_paths:
        try:
            if os.path.exists(path):
                os.remove(path)
        except:
            pass

# Data Processing
def create_client(row, client_field, contact_field='Contact'):
    """Create or get client from row data."""
    client_name = row.get(client_field, 'Unknown') if not pd.isna(row.get(client_field)) else 'Unknown'
    contact = row.get(contact_field, '').split(';')[0] if not pd.isna(row.get(contact_field, '')) else ''
    client, _ = Client.objects.get_or_create(
        name=client_name,
        defaults={
            'contact_info': contact,
            'phone': '',
            'email': contact if '@' in contact else ''
        }
    )
    return client

def create_building(row, client, address_field='Building', bbl_field='Block'):
    """Create building from row data."""
    return Building.objects.create(
        address=row.get(address_field, 'Unknown') if not pd.isna(row.get(address_field)) else 'Unknown',
        borough='Unknown',
        BBL=str(row.get(bbl_field, '')) if not pd.isna(row.get(bbl_field)) else '',
        BIN=str(row.get('BIN', '')) if not pd.isna(row.get('BIN')) else '',
        number_of_bins=int(row.get('# BINs', 0)) if not pd.isna(row.get('# BINs')) else 0,
        client=client
    )

def process_ll84_row(row, building, service):
    """Process LL84 tracking data for a building."""
    tracking_data = {
        col: str(row[col]) if not pd.isna(row[col]) else 'Unknown'
        for col in row.index
        if col not in ['Building Address', '2024 File', 'Contact']
    }
    BuildingTracking.objects.get_or_create(
        building=building,
        service=service,
        defaults={'tracking_info': json.dumps(tracking_data, ensure_ascii=False)}
    )

# Views
@login_required
def home(request):
    """Render home page with buildings and clients filtered by query."""
    query = request.GET.get('q', '')
    tab = request.GET.get('tab', 'buildings')
    
    buildings = Building.objects.prefetch_related('services', 'client').all()
    clients = Client.objects.all()
    
    if query:
        buildings = buildings.filter(
            Q(address__icontains=query) |
            Q(BBL__icontains=query) |
            Q(BIN__icontains=query) |
            Q(client__name__icontains=query)
        )
        if tab == 'clients':
            clients = clients.filter(
                Q(name__icontains=query) |
                Q(phone__icontains=query) |
                Q(email__icontains=query)
            )
    
    context = {
        'tab': tab,
        'query': query,
        'buildings': buildings,
        'clients': clients
    }
    return render(request, 'home.html', context)

@login_required
def building_detail(request, building_id):
    """Display building details."""
    building = get_object_or_404(Building, id=building_id)
    context = {
        'building': building,
        'services': building.services.all(),
        'client': building.client,
        'trackings': building.trackings.all()
    }
    return render(request, 'building_detail.html', context)  # Changed from '-LL84-Benchmarking'

def process_excel_import(request, sheet_type, excel_file):
    """Process Excel file import based on sheet type."""
    temp_file_path = save_temp_excel(excel_file)
    reformatted_file_path = None
    
    try:
        if sheet_type == 'LL84-Benchmarking':
            reformatted_file_path = temp_file_path.replace('.xlsx', '_reformatted.xlsx')
            workbook = openpyxl.load_workbook(temp_file_path)
            workbook = unmerge_excel_cells(workbook)
            workbook.save(reformatted_file_path)
            excel_data = pd.ExcelFile(reformatted_file_path)
            
            if 'LL84- Benchmarking' not in excel_data.sheet_names:
                raise ValueError("Excel file does not contain 'LL84- Benchmarking' tab")
            
            df = pd.read_excel(excel_data, sheet_name='LL84- Benchmarking')
            df.columns = df.columns.str.strip()
            ll84_service, _ = Service.objects.get_or_create(name='LL84')
            
            for _, row in df.iterrows():
                if row.get('Building', '').lower() == 'archive buildings below:':
                    break
                client = create_client(row, 'Customer')
                building = create_building(row, client)
                building.services.add(ll84_service)
                process_ll84_row(row, building, ll84_service)
                
        elif sheet_type == 'LL97':
            excel_data = pd.ExcelFile(temp_file_path)
            df = pd.read_excel(excel_data, sheet_name='LL97')
            df.columns = df.columns.str.strip()
            service, _ = Service.objects.get_or_create(name='LL97')
            
            for _, row in df.iterrows():
                client = create_client(row, '2024 File')
                building = create_building(row, client, 'Building Address', 'BBL')
                if not pd.isna(row.get('LL97-321')):
                    building.services.add(service)
                    
        else:
            raise ValueError('Invalid sheet type selected.')
            
        messages.success(request, f'Imported data from {sheet_type} successfully!')
        
    except Exception as e:
        messages.error(request, f'Error importing file: {str(e)}')
        
    finally:
        cleanup_temp_files(temp_file_path, reformatted_file_path)
    
    return redirect('home')

@login_required
def import_excel(request):
    """Handle Excel file upload and processing."""
    if request.method != 'POST':
        return render(request, 'home.html')
        
    excel_file = request.FILES.get('excel_file')
    sheet_type = request.POST.get('sheet_type')
    
    if not excel_file:
        messages.error(request, 'No file uploaded.')
        return redirect('home')
    if not sheet_type:
        messages.error(request, 'Please select a sheet type.')
        return redirect('home')
        
    return process_excel_import(request, sheet_type, excel_file)

# Google Sheets Integration
def get_gsheet_client():
    """Initialize Google Sheets client."""
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name('credentials/sfe-import-credentials.json', scope)
    return gspread.authorize(creds)

@login_required
def preview_gsheet(request):
    """Preview Google Sheet data."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method.'}, status=405)
        
    try:
        data = json.loads(request.body)
        sheet_id = data.get('sheet_id')
        if not sheet_id:
            return JsonResponse({'error': 'Please provide a valid Google Sheet ID.'}, status=400)
            
        client = get_gsheet_client()
        sheet = client.open_by_key(sheet_id).sheet1
        data = sheet.get_all_records()
        return JsonResponse({'sheet_id': sheet_id, 'preview': data}, status=200)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def import_gsheet(request):
    """Import data from Google Sheet."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method.'}, status=405)
        
    try:
        data = json.loads(request.body)
        sheet_id = data.get('sheet_id')
        if not sheet_id:
            return JsonResponse({'error': 'Please provide a valid Google Sheet ID.'}, status=400)
            
        client = get_gsheet_client()
        sheet = client.open_by_key(sheet_id).sheet1
        data = sheet.get_all_records()
        
        ll97_service, _ = Service.objects.get_or_create(name='LL97')
        ll88_service, _ = Service.objects.get_or_create(name='LL88')
        
        for row in data:
            client = create_client(row, '2024 File', 'Contact Email')
            building = create_building(row, client, 'Building Address', 'BBL')
            if row.get('LL97 Schedule'):
                building.services.add(ll97_service)
            if row.get('LL88 Needed') == 'Y':
                building.services.add(ll88_service)
                
        return JsonResponse({'message': 'Data imported successfully.'}, status=200)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Admin Views
@login_required
@user_passes_test(is_admin_user)
def admin_view(request):
    """Display admin dashboard with statistics."""
    context = {
        'buildings': Building.objects.prefetch_related('services', 'client').all(),
        'building_count': Building.objects.count(),
        'client_count': Client.objects.count(),
        'service_count': Service.objects.count(),
        'tracking_count': BuildingTracking.objects.count(),
    }
    return render(request, 'admin_view.html', context)

@login_required
@user_passes_test(is_admin_user)
def delete_all_data(request):
    """Delete all buildings, clients, and services."""
    if request.method != 'POST':
        return redirect('home')
        
    try:
        Building.objects.all().delete()
        Client.objects.all().delete()
        Service.objects.all().delete()
        messages.success(request, "All data deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting data: {str(e)}")
        
    return redirect(reverse('home') + '?tab=buildings')

# Tracking Updates
FIELD_MAPPINGS = {
    'Price': ('LL84_Price', lambda x: str(x)),
    '2020 Filed': ('LL84_2020_Filed', lambda x: x.strip().lower() == 'y'),
    '2021PMT': ('LL84_2021PMT', lambda x: str(x)),
    '2022 Score to customer': ('LL84_2022_Score_to_customer', lambda x: x.strip().lower() == 'y'),
    '2023 Invoice sent': ('LL84_2023_Invoice_sent', lambda x: x.strip().lower() == 'y'),
    '2023 PMT': ('LL84_2023_PMT', lambda x: x.strip().lower() == 'y'),
    'Submission': ('LL84_Submission', lambda x: x.strip().lower() == 'y'),
    '23 Grading': ('LL84_23_Grading', lambda x: x.strip().lower() == 'y'),
    '24 Outreach': ('LL84_24_Outreach', lambda x: x.strip().lower() == 'y'),
    '2024 PMT': ('LL84_2024_PMT', lambda x: x.strip().lower() == 'y'),
    '24 Submission': ('LL84_24_Submission', lambda x: x.strip().lower() == 'y'),
    '2024 Score': ('LL84_2024_Score', lambda x: str(x)),
    'water benchmark': ('LL84_water_benchmark', lambda x: x.strip().lower() == 'y'),
    '2025 Paid': ('LL84_2025_Paid', lambda x: x.strip().lower() == 'y'),
    '2025 submission': ('LL84_2025_submission', lambda x: str(x)),
    'Confirmation Email': ('LL84_Confirmation_Email', lambda x: str(x)),
    'Show data in BEAM': ('LL84_Show_data_in_BEAM', lambda x: str(x)),
}

def update_tracking_from_json(tracking, tracking_info):
    """Update tracking fields from JSON data."""
    for json_key, (field_name, transform) in FIELD_MAPPINGS.items():
        if json_key in tracking_info:
            setattr(tracking, field_name, transform(tracking_info[json_key]))
    tracking.save()

@login_required
@user_passes_test(is_admin_user)
def building_update_bins(request, building_id):
    """Update LL84 tracking data for a single building."""
    building = get_object_or_404(Building, id=building_id)
    
    try:
        tracking = BuildingTracking.objects.get(building=building, service__name='LL84')
        tracking_info = json.loads(tracking.tracking_info)
        update_tracking_from_json(tracking, tracking_info)
        messages.success(request, f'LL84 tracking data for "{building.address}" updated successfully.')
    except BuildingTracking.DoesNotExist:
        messages.error(request, f'No LL84 tracking data found for "{building.address}".')
    except json.JSONDecodeError:
        messages.error(request, f'Invalid tracking_info format for "{building.address}".')
    except Exception as e:
        messages.error(request, f'Error updating tracking data: {str(e)}')
        
    return redirect('admin_view')

@login_required
@user_passes_test(is_admin_user)
def process_all_buildings(request):
    """Update LL84 tracking data for all buildings."""
    updated_count = 0
    error_count = 0
    
    for building in Building.objects.all():
        try:
            tracking = BuildingTracking.objects.get(building=building, service__name='LL84')
            tracking_info = json.loads(tracking.tracking_info)
            update_tracking_from_json(tracking, tracking_info)
            updated_count += 1
        except BuildingTracking.DoesNotExist:
            logger.warning(f'No LL84 tracking data found for "{building.address}".')
            error_count += 1
        except json.JSONDecodeError:
            logger.error(f'Invalid tracking_info format for "{building.address}".')
            error_count += 1
        except Exception as e:
            logger.error(f'Error updating tracking data for "{building.address}": {str(e)}')
            error_count += 1
    
    if updated_count > 0:
        messages.success(request, f'Successfully updated LL84 tracking data for {updated_count} building(s).')
    if error_count > 0:
        messages.error(request, f'Failed to update {error_count} building(s) due to missing or invalid data.')
    
    return redirect('admin_view')

@login_required
@user_passes_test(is_admin_user)
def update_tracking(request):
    """Update individual tracking record."""
    if request.method != 'POST':
        return redirect('home')
        
    tracking_id = request.POST.get('tracking_id')
    tracking = get_object_or_404(BuildingTracking, id=tracking_id)
    
    try:
        fields = {
            'LL84_Price': request.POST.get('LL84_Price', tracking.LL84_Price),
            'LL84_2020_Filed': request.POST.get('LL84_2020_Filed') == 'Yes',
            'LL84_2021PMT': request.POST.get('LL84_2021PMT', tracking.LL84_2021PMT),
            'LL84_2022_Score_to_customer': request.POST.get('LL84_2022_Score_to_customer') == 'Yes',
            'LL84_2023_Invoice_sent': request.POST.get('LL84_2023_Invoice_sent') == 'Yes',
            'LL84_2023_PMT': request.POST.get('LL84_2023_PMT') == 'Yes',
            'LL84_Submission': request.POST.get('LL84_Submission') == 'Yes',
            'LL84_23_Grading': request.POST.get('LL84_23_Grading') == 'Yes',
            'LL84_24_Outreach': request.POST.get('LL84_24_Outreach') == 'Yes',
            'LL84_2024_PMT': request.POST.get('LL84_2024_PMT') == 'Yes',
            'LL84_24_Submission': request.POST.get('LL84_24_Submission') == 'Yes',
            'LL84_2024_Score': request.POST.get('LL84_2024_Score', tracking.LL84_2024_Score),
            'LL84_water_benchmark': request.POST.get('LL84_water_benchmark') == 'Yes',
            'LL84_2025_Paid': request.POST.get('LL84_2025_Paid') == 'Yes',
            'LL84_2025_submission': request.POST.get('LL84_2025_submission', tracking.LL84_2025_submission),
            'LL84_Confirmation_Email': request.POST.get('LL84_Confirmation_Email', tracking.LL84_Confirmation_Email),
            'LL84_Show_data_in_BEAM': request.POST.get('LL84_Show_data_in_BEAM', tracking.LL84_Show_data_in_BEAM),
        }
        
        for field, value in fields.items():
            setattr(tracking, field, value)
        
        tracking.save()
        messages.success(request, f'LL84 tracking data for "{tracking.building.address}" updated successfully.')
    except Exception as e:
        messages.error(request, f'Error updating tracking data: {str(e)}')
        
    return redirect('building_detail', building_id=tracking.building.id)


@login_required
def client_detail(request, client_id):
    """Display client details and associated buildings."""
    client = get_object_or_404(Client, id=client_id)
    buildings = Building.objects.filter(client=client)
    context = {
        'client': client,
        'buildings': buildings
    }
    return render(request, 'client_detail.html', context)